# 导入所需的库
import openpyxl  # 用于处理 Excel 文件
from selenium import webdriver  # 用于模拟浏览器操作
from selenium.webdriver.common.by import By  # 用于定位网页元素
import time  # 用于添加延迟等待
import requests  # 用于发送 HTTP 请求
import parsel  # 用于解析 HTML 数据
import pandas as pd  # 用于处理数据和生成 Excel 文件

# 使用 Selenium webdriver 打开 Chrome 浏览器
driver = webdriver.Chrome()

# 打开链家网站首页
url = 'https://bj.lianjia.com/'
driver.get(url)
driver.maximize_window()
time.sleep(5)

# 点击“二手房”对应的链接
xiaoqu_link = driver.find_element(By.XPATH,'//li[@data-click-evtid="20599" and contains(@data-action, "click_name=二手房")]')
xiaoqu_link.click()
time.sleep(10)

# 切换到新打开的窗口
driver.switch_to.window(driver.window_handles[1])
current_url = driver.current_url

# 创建一个新的 Excel 文件
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# 设置 Excel 表头
sheet['A1'] = '二手房'
sheet['B1'] = '链接'

# 逐个点击链接，并将 href 和 text 写入 Excel
row = 2  # 从第二行开始写入数据

# 获取当前页面中的小区链接列表
links = driver.find_elements(By.XPATH, '//div[@class="info clear"]//a')

# 遍历链接列表，获取文本和链接，并写入 Excel 文件
for link in links:
    text = link.get_attribute("text")
    href = link.get_attribute("href")
    print(text)
    print(href)
    # 检查 text 和 href 是否存在
    if text and href:
        # 写入数据到 Excel 文件中
        sheet[f'A{row}'] = text
        sheet[f'B{row}'] = href
        row += 1

time.sleep(40)

# 遍历多个页面
for page in range(2, 100):
    # 更新 URL 以包含页码
    page_url = f"{current_url}pg{page}/"
    driver.get(page_url)
    print(page_url)
    links = driver.find_elements(By.XPATH, '//div[@class="info clear"]//a')
    for link in links:
        text = link.get_attribute("text")
        href = link.get_attribute("href")

        # 检查 text 和 href 是否存在
        if text and href:
            # 写入数据到 Excel 文件中
            sheet[f'A{row}'] = text
            sheet[f'B{row}'] = href
            row += 1
    time.sleep(40)

# 保存 Excel 文件
wb.save('2-二手房链 接.xlsx')
time.sleep(10)

# 读取之前保存的 Excel 文件
wb = openpyxl.load_workbook('2-二手房链接.xlsx')
sheet = wb['Sheet1']
result_list = []

# 遍历 Excel 文件中的链接
for row in range(2, sheet.max_row + 1, 3):
    href = sheet[f'B{row}'].value
    if href:
        try:
            # 获取链接内容并解析
            driver.get(href)
            time.sleep(10)
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36'
            }

            response = requests.get(url=href, headers=headers)
            content = response.content.decode('utf-8')
            selector = parsel.Selector(response.text)
            time.sleep(5)

            # 提取相关信息
            community1 = selector.css('.introContent.showbasemore div')
            community = community1.css('.baseattribute.clear div::text').getall()
            print(community)
            division = community[1]
            address = community[3]
            address1 = community[5]

            lis = selector.css('.aroundInfo div')
            sale = lis.css('.communityName a::text').getall()
            division1 = sale[0]

            dit = {
                '小区名称': division1,
                '核心卖点': division,
                '小区介绍': address,
                '周边配套': address1,
            }
            result_list.append(dit)
            print(dit)

            # 将数据存入 DataFrame 并保存到 Excel 文件中
            df = pd.DataFrame(result_list)
            df.to_excel('2-北京小区简介数据.xlsx', index=False)
        except Exception as e:
            print(f"Error processing URL {href}: {e}")
            continue
    time.sleep(5)
