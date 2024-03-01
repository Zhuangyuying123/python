import urllib  # 用于处理URL
import openpyxl  # 用于Excel文件操作
import parsel as parsel  # 网页解析库
from selenium import webdriver  # 浏览器自动化工具
from selenium.webdriver.common.by import By  # 定位元素方法
import time  # 时间处理模块
import requests  # 发送HTTP请求模块
import pandas as pd  # 数据处理库

# 使用Chrome浏览器作为webdriver
driver = webdriver.Chrome()

# 打开链家网站首页
url = 'https://bj.lianjia.com/'
driver.get(url)
driver.maximize_window()
time.sleep(5)  # 等待页面加载

# 点击“小区”对应的链接
xiaoqu_link = driver.find_element(By.XPATH, '//li[@data-click-evtid="20599" and contains(@data-action, "click_name=小区")]')
xiaoqu_link.click()
time.sleep(10)

# 切换到新打开的窗口
driver.switch_to.window(driver.window_handles[1])
current_url = driver.current_url

# 基于xpath定位，获取页面中的小区链接列表
links = driver.find_elements(By.XPATH, '//div[@data-role="ershoufang"]//a')

# 创建一个新的 Excel 文件
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# 设置 Excel 表头
sheet['A1'] = '区划'
sheet['B1'] = '链接'

# 逐个点击链接，并将 href 和 text 写入 Excel
row = 2  # 从第二行开始写入数据
for link in links:
    text = link.get_attribute("text")  # 获取链接文本
    href = link.get_attribute("href")  # 获取链接地址

    # 检查 text 和 href 是否存在
    if text and href:
        # 写入数据到 Excel 文件中
        sheet[f'A{row}'] = text
        sheet[f'B{row}'] = href
        row += 1

# 保存 Excel 文件
wb.save('1-各区划链接.xlsx')

time.sleep(10)

# 加载保存的Excel文件
wb = openpyxl.load_workbook('1-各区划链接.xlsx')
sheet = wb['Sheet1']
result_list = []

# 遍历 Excel 表中的链接
for row in range(2, sheet.max_row + 1):
    href = sheet[f'B{row}'].value  # 获取Excel中的链接
    if href:
        driver.get(href)  # 访问链接
        print(href)
        time.sleep(20)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36'
        }
        # 获取当前页面的完整URL
        current_url = driver.current_url
        # 如果href是相对链接，将其与基本URL拼接
        if not href.startswith('http'):
            href = urllib.parse.urljoin(current_url, href)
        # 发送请求并解析页面内容
        response = requests.get(url=href, headers=headers)
        content = response.content.decode('utf-8')
        selector = parsel.Selector(response.text)
        time.sleep(20)
        lis = selector.css('.listContent li')
        # 提取小区信息并保存到列表
        for li in lis:
            # 小区名称
            community_name = li.css('.title a::text').get()
            # 销售情况
            sale = li.css('.houseInfo a::text').getall()
            # 地址
            positionInfo = li.css('.positionInfo a::text').getall()
            # 区划
            division = positionInfo[0]
            # 区域
            address = positionInfo[1]
            # 材质及年份
            material = li.css('.positionInfo::text').getall()
            material_str = ''.join(material).replace('\n', '').replace(' ', '').replace('\xa0', '')
            material_list = material_str.split('/')
            building_type = '/'.join(material_list[:-1])  # 获取除最后一个元素外的所有元素
            year_built = material_list[-1]  # 获取最后一个元素
            # 周边地铁
            subway = li.css('.tagList span::text').get()
            # 房价
            Price = li.css('.totalPrice span::text').get() + '万'
            # 在售房数
            housecount = li.css('.xiaoquListItemSellCount a span::text').get() + '套'
            dit = {
                '小区名称': community_name,
                '销售情况': sale,
                '区划': division,
                '区域': address,
                '周边地铁': subway,
                '建筑材质': building_type,
                '建造年份': year_built,
                '房价': Price,
                '在售房数': housecount,
            }
            result_list.append(dit)
            print(dit)

        for page in range(2, 30):
            # 更新URL以包含页码
            page_url = f"{href}pg{page}/"
            driver.get(page_url)
            print(page_url)
            time.sleep(20)
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36'
            }
            # 获取当前页面的完整URL
            current_url = driver.current_url
            # 如果href是相对链接，将其与基本URL拼接
            if not page_url.startswith('http'):
                page_url = urllib.parse.urljoin(current_url, page_url)
            response = requests.get(url=page_url, headers=headers)
            content = response.content.decode('utf-8')
            selector = parsel.Selector(response.text)
            lis = selector.css('.listContent li')
            for li in lis:
                # 标题
                community_name = li.css('.title a::text').get()
                # 出售情况
                sale = li.css('.houseInfo a::text').getall()
                # 地址
                positionInfo = li.css('.positionInfo a::text').getall()
                # 区划
                division = positionInfo[0]
                # 地名
                address = positionInfo[1]
                # 材质及年份
                material = li.css('.positionInfo::text').getall()
                material_str = ''.join(material).replace('\n', '').replace(' ', '').replace('\xa0', '')
                material_list = material_str.split('/')
                building_type = '/'.join(material_list[:-1])  # 获取除最后一个元素外的所有元素
                year_built = material_list[-1]  # 获取最后一个元素
                # 周边地铁
                subway = li.css('.tagList span::text').get()
                # 房价
                Price = li.css('.totalPrice span::text').get() + '万'
                # 房子套数
                housecount = li.css('.xiaoquListItemSellCount a span::text').get() + '套'
                dit = {
                    '小区名称': community_name,
                    '销售情况': sale,
                    '区划': division,
                    '区域': address,
                    '周边地铁': subway,
                    '建筑材质': building_type,
                    '建造年份': year_built,
                    '房价': Price,
                    '在售房数': housecount,
                }
                result_list.append(dit)
                print(dit)
                # 将结果列表转换为DataFrame并保存到Excel文件中
                df = pd.DataFrame(result_list)
                df.to_excel('1-北京小区数据.xlsx', index=False)
    time.sleep(50)
